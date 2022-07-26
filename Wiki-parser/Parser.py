import openpyxl
import requests
from bs4 import BeautifulSoup

book = openpyxl.Workbook()
sheet = book.active
sheet['A1'] = 'Город'
sheet['B1'] = 'Регион'
sheet['C1'] = 'Федеральный округ'
sheet['D1'] = 'Население'
sheet['E1'] = 'Основание или первое упоминание'
sheet['F1'] = 'Статус города'


def parser():
    URL = 'https://ru.wikipedia.org/wiki/Список_городов_России'
    res = requests.get(URL).text
    soup = BeautifulSoup(res, "html.parser")
    for counter in range(2, 1119):
        tables = soup.find("table").find_all('tr')[counter]

        city = tables.find_all('td')[2].text
        a1 = sheet.cell(row=counter, column=1)
        a1.value = city

        region = tables.find_all('td')[3].text
        b1 = sheet.cell(row=counter, column=2)
        b1.value = region

        dist = tables.find_all('td')[4].text
        c1 = sheet.cell(row=counter, column=3)
        c1.value = dist

        population = tables.find_all('td')[5].text.replace(" ", "")
        d1 = sheet.cell(row=counter, column=4)
        d1.value = int(population)

        try:
            year = tables.find_all('td')[6].text.replace(" ", "")
            e1 = sheet.cell(row=counter, column=5)
            e1.value = int(year)
        except:
            e1.value = year

        try:
            status = tables.find_all('td')[7].text.replace(" ", "")
            f1 = sheet.cell(row=counter, column=6)
            f1.value = int(status)
        except:
            f1.value = year

    book.save('book.xlsx')
    book.close()


if __name__ == "__main__":
    parser()
